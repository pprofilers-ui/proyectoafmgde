import argparse
import os
import re
import sys
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))


@dataclass
class ChamberLocationSeed:
    code: str
    name: str
    room: str
    shelf: str
    position: str


@dataclass
class ChamberSeed:
    code: str
    name: str
    location: str
    condition_code: str
    condition_name: str
    temperature: Decimal
    humidity: Decimal | None
    locations: list[ChamberLocationSeed]


def normalize_spaces(value: str) -> str:
    return re.sub(r"\s+", " ", (value or "").strip())


def parse_condition(raw_value: str) -> tuple[Decimal, Decimal | None]:
    cleaned = normalize_spaces(raw_value).replace(",", ".")
    numbers = re.findall(r"\d+(?:\.\d+)?", cleaned)
    if not numbers:
        raise ValueError(f"No se pudo interpretar la condición: {raw_value!r}")
    temperature = Decimal(numbers[0])
    humidity = Decimal(numbers[1]) if len(numbers) > 1 else None
    return temperature, humidity


def build_condition_code(temperature: Decimal, humidity: Decimal | None) -> str:
    temp_part = f"{int(temperature):02d}"
    humidity_part = f"{int(humidity):02d}" if humidity is not None else "NA"
    return f"COND-{temp_part}-{humidity_part}"


def build_condition_name(temperature: Decimal, humidity: Decimal | None) -> str:
    if humidity is None:
        return f"{int(temperature)}C"
    return f"{int(temperature)}/{int(humidity)}"


def extract_chambers_from_workbook(excel_path: Path) -> list[ChamberSeed]:
    workbook = load_workbook(excel_path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))

    header_row = rows[2]
    subheader_row = rows[3]
    data_rows = rows[4:]

    chamber_columns = []
    for index, value in enumerate(header_row):
        label = normalize_spaces(str(value or ""))
        if label.upper().startswith("CÁMARA") or label.upper().startswith("CAMARA"):
            chamber_columns.append((index, label))

    chambers: list[ChamberSeed] = []
    for chamber_number, (base_col, label) in enumerate(chamber_columns, start=1):
        condition_value = normalize_spaces(str(data_rows[0][base_col] or ""))
        if not condition_value:
            continue

        temperature, humidity = parse_condition(condition_value)
        condition_code = build_condition_code(temperature, humidity)
        condition_name = build_condition_name(temperature, humidity)

        locations: list[ChamberLocationSeed] = []
        for row in data_rows:
            shelf_value = normalize_spaces(str(row[base_col + 1] or ""))
            position_value = normalize_spaces(str(row[base_col + 2] or ""))
            if not shelf_value or not position_value:
                continue
            if shelf_value.upper().startswith("ESTANTERÍA SALA"):
                continue

            locations.append(
                ChamberLocationSeed(
                    code=f"CAM{chamber_number:02d}-{shelf_value}",
                    name=f"Cámara {chamber_number} - Estantería {shelf_value}",
                    room=f"Cámara {chamber_number}",
                    shelf=shelf_value,
                    position=position_value,
                )
            )

        chambers.append(
            ChamberSeed(
                code=f"CAM-{chamber_number:02d}",
                name=f"Cámara {chamber_number}",
                location="Sala de cámaras",
                condition_code=condition_code,
                condition_name=condition_name,
                temperature=temperature,
                humidity=humidity,
                locations=locations,
            )
        )

    if not chambers:
        raise ValueError("No se encontraron cámaras válidas en el Excel.")
    return chambers


def print_plan(chambers: list[ChamberSeed]) -> None:
    print("Cámaras detectadas en el Excel:")
    for chamber in chambers:
        humidity_text = f" / {int(chamber.humidity)}% HR" if chamber.humidity is not None else ""
        print(
            f"- {chamber.code} | {chamber.name} | {int(chamber.temperature)}C{humidity_text} | "
            f"ubicaciones: {len(chamber.locations)}"
        )


def apply_reset(excel_path: Path) -> None:
    os.environ.setdefault("DJANGO_SETTINGS_MODULE", "config.settings")
    import django

    django.setup()

    from django.db import transaction
    from stability.models import Chamber, ChamberDeviation, ChamberLocation, Sample, SampleSchedule, StorageCondition

    chambers = extract_chambers_from_workbook(excel_path)
    print_plan(chambers)

    with transaction.atomic():
        sample_updates = Sample.objects.exclude(chamber__isnull=True).update(
            chamber=None,
            shelf="",
            tray="",
            container="",
            physical_position="",
        )
        schedule_updates = SampleSchedule.objects.filter(
            chamber__isnull=False
        ).update(
            chamber=None,
            chamber_location=None,
        )
        deviation_deletes = ChamberDeviation.objects.count()

        ChamberDeviation.objects.all().delete()
        Chamber.objects.all().delete()
        ChamberLocation.objects.all().delete()
        StorageCondition.objects.all().delete()

        conditions_by_code = {}
        for chamber in chambers:
            if chamber.condition_code not in conditions_by_code:
                condition = StorageCondition.objects.create(
                    code=chamber.condition_code,
                    name=chamber.condition_name,
                    temperature_set_point=chamber.temperature,
                    humidity_set_point=chamber.humidity,
                    light_condition="",
                    is_active=True,
                )
                conditions_by_code[chamber.condition_code] = condition

        for chamber in chambers:
            created_locations = []
            for location in chamber.locations:
                created_locations.append(
                    ChamberLocation.objects.create(
                        code=location.code,
                        name=location.name,
                        room=location.room,
                        shelf=location.shelf,
                        position=location.position,
                        is_active=True,
                    )
                )

            Chamber.objects.create(
                code=chamber.code,
                name=chamber.name,
                location=chamber.location,
                storage_condition=conditions_by_code[chamber.condition_code],
                chamber_location=created_locations[0] if created_locations else None,
                temperature_set_point=chamber.temperature,
                humidity_set_point=chamber.humidity,
                is_active=True,
            )

    print("")
    print("Sincronización aplicada.")
    print(f"- Muestras desvinculadas de cámara: {sample_updates}")
    print(f"- Fechas de muestreo desvinculadas de cámara/ubicación: {schedule_updates}")
    print(f"- Desviaciones eliminadas: {deviation_deletes}")
    print(f"- Cámaras creadas: {len(chambers)}")
    print(f"- Ubicaciones creadas: {sum(len(chamber.locations) for chamber in chambers)}")
    print(f"- Condiciones creadas: {len({chamber.condition_code for chamber in chambers})}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Reinicia cámaras, ubicaciones y condiciones de almacenamiento a partir de un Excel."
    )
    parser.add_argument("excel_path", help="Ruta al Excel de cámaras.")
    parser.add_argument(
        "--apply",
        action="store_true",
        help="Aplica los cambios. Sin esta opción solo muestra el plan.",
    )
    args = parser.parse_args()

    excel_path = Path(args.excel_path).expanduser().resolve()
    if not excel_path.exists():
        raise FileNotFoundError(f"No existe el Excel: {excel_path}")

    chambers = extract_chambers_from_workbook(excel_path)
    print_plan(chambers)

    if not args.apply:
        print("")
        print("Dry run completado. Usa --apply para ejecutar la limpieza y recreación.")
        return

    apply_reset(excel_path)


if __name__ == "__main__":
    main()
